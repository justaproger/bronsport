# --- START OF FULL FILE backend/dashboard/views.py ---
from django.shortcuts import render, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from django.utils.translation import gettext_lazy as _, gettext
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg 
from django.db.models.functions import Coalesce, TruncDay, TruncMonth
from decimal import Decimal, ROUND_HALF_UP
from datetime import timedelta, date as dt_date, datetime, time 
import json 
import logging
from .models import StatisticsDashboard # <--- ВОТ ДОБАВЛЕННЫЙ ИМПОРТ

from django.contrib import admin, messages
from django.contrib.auth import get_user_model
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from core.utils import get_admin_university_details

try:
    from universities.models import University, Staff, SportClub
    UNIVERSITY_APP_AVAILABLE = True
except ImportError:
    University = Staff = SportClub = None # type: ignore
    UNIVERSITY_APP_AVAILABLE = False
try:
    from facilities.models import Facility
    FACILITY_APP_AVAILABLE = True
except ImportError:
    Facility = None # type: ignore
    FACILITY_APP_AVAILABLE = False
try:
    from bookings.models import Order 
    BOOKING_APP_AVAILABLE = True
except ImportError:
     Order = None # type: ignore
     BOOKING_APP_AVAILABLE = False

User = get_user_model()
logger = logging.getLogger(__name__)

@staff_member_required
def statistics_dashboard_view(request):
    user = request.user
    admin_university_obj = get_admin_university_details(user)
    
    # Базовый контекст
    base_context = {
        'title': _('Статистика'), 
        'site_title': getattr(admin.site, 'site_title', _('Django site admin')),
        'site_header': getattr(admin.site, 'site_header', _('Django administration')),
        'has_permission': True, 
        'user': user, 
        'is_superuser': user.is_superuser,
        'opts': StatisticsDashboard._meta, # Для хлебных крошек и заголовка в admin base
        'app_label': StatisticsDashboard._meta.app_label, # Для хлебных крошек
    }

    context = {
        **base_context,
        'university_context': None,
        'stats_kpi_cards': {}, # Изменено для структурированных KPI
        'universities_list_for_select': None,
        'universities_stats_table': None,
        'current_selected_uni_id': None,
    }
    
    target_university_for_stats = None 
    is_global_view_active = False

    if user.is_superuser:
        uni_id_str = request.GET.get('university_id')
        context['current_selected_uni_id'] = uni_id_str # Сохраняем для select'а
        if uni_id_str and uni_id_str != "global" and University: # "global" для общей статистики
            try: 
                target_university_for_stats = University.objects.get(pk=int(uni_id_str), is_active=True)
                context['title'] = _('Статистика ВУЗа: {uni_name}').format(uni_name=target_university_for_stats.name)
            except (University.DoesNotExist, ValueError, TypeError): 
                messages.error(request, _("Выбранный университет не найден или неактивен. Показывается общая статистика."))
                target_university_for_stats = None; is_global_view_active = True
                context['title'] = _('Общая Статистика по Платформе')
        else: 
            is_global_view_active = True
            context['title'] = _('Общая Статистика по Платформе')
            context['current_selected_uni_id'] = "global" # Устанавливаем значение для "Общей статистики"
        
        if University: # Список ВУЗов для выбора суперюзером
            context['universities_list_for_select'] = University.objects.filter(is_active=True).order_by('name')
    elif admin_university_obj and admin_university_obj != -1:
        target_university_for_stats = admin_university_obj
        context['title'] = _('Статистика ВУЗа: {uni_name}').format(uni_name=target_university_for_stats.name)
    else: 
        messages.error(request, _("У вас нет прав или назначенного университета для просмотра статистики."))
        return render(request, 'admin/dashboard/statistics_dashboard.html', context)
        
    context['university_context'] = target_university_for_stats

    if not (Order and Facility and User and BOOKING_APP_AVAILABLE):
        error_msg = _("Необходимые модели для статистики не загружены. Проверьте конфигурацию приложений.")
        messages.error(request, error_msg)
        # Можно добавить error_msg в context, если шаблон его отображает
        return render(request, 'admin/dashboard/statistics_dashboard.html', context)

    # Сбор статистики
    if target_university_for_stats or is_global_view_active:
        now = timezone.localtime(timezone.now())
        today_date = now.date()
        start_of_this_month = today_date.replace(day=1)
        start_of_this_week = today_date - timedelta(days=today_date.weekday())
        
        base_orders_qs = Order.objects.all()
        if target_university_for_stats: # Если выбран конкретный ВУЗ
            base_orders_qs = base_orders_qs.filter(facility__university=target_university_for_stats)
        
        successful_statuses = [Order.STATUS_CONFIRMED, Order.STATUS_COMPLETED]
        all_orders_ever_successful_qs = base_orders_qs.filter(status__in=successful_statuses)

        # --- KPI Карточки ---
        kpi_cards = {}

        # Заказы (Этот месяц)
        orders_this_month_qs = all_orders_ever_successful_qs.filter(created_at__date__gte=start_of_this_month, created_at__date__lte=today_date)
        month_agg = orders_this_month_qs.aggregate(total_count=Count('id'), total_revenue=Coalesce(Sum('total_price'), Decimal(0)))
        kpi_cards['month_orders_count'] = {
            'label': _('Заказов (Этот месяц)'),
            'value': month_agg['total_count'],
        }
        month_subscriptions_revenue = orders_this_month_qs.filter(order_type=Order.TYPE_SUBSCRIPTION).aggregate(total=Coalesce(Sum('total_price'), Decimal(0)))['total']
        kpi_cards['month_revenue'] = {
            'label': _('Доход (Этот месяц)'),
            'value': f"{month_agg['total_revenue']:,} {gettext('сум')}".replace(",", " "),
            'sub_text': _('Вкл. подписки: {amount_sub} {currency}').format(amount_sub=f"{month_subscriptions_revenue:,}".replace(",", " "), currency=gettext('сум'))
        }

        # Заказы (Эта неделя)
        orders_this_week_qs = all_orders_ever_successful_qs.filter(created_at__date__gte=start_of_this_week, created_at__date__lte=today_date)
        week_agg = orders_this_week_qs.aggregate(total_count=Count('id'),total_revenue=Coalesce(Sum('total_price'), Decimal(0)))
        kpi_cards['week_orders_count'] = {
            'label': _('Заказов (Эта неделя)'),
            'value': week_agg['total_count'],
        }
        kpi_cards['week_revenue'] = {
            'label': _('Доход (Эта неделя)'),
            'value': f"{week_agg['total_revenue']:,} {gettext('сум')}".replace(",", " "),
        }
        
        # Активные заказы сегодня
        # Логика для active_orders_today_count (нужно уточнить, как в MyUnifiedOrderListView)
        q_bookings_active_today = Q(order_type__in=[Order.TYPE_SLOT_BOOKING, Order.TYPE_ENTRY_FEE]) & \
                                  Q(status=Order.STATUS_CONFIRMED) & Q(booking_date=today_date) & \
                                  Q(facility__close_time__gt=now.time())
        q_subscriptions_active_today = Q(order_type=Order.TYPE_SUBSCRIPTION) & \
                                       Q(status=Order.STATUS_CONFIRMED) & \
                                       Q(subscription_start_date__lte=today_date) & \
                                       Q(subscription_end_date__gte=today_date)
        active_orders_today_count = base_orders_qs.filter(q_bookings_active_today | q_subscriptions_active_today).count()
        kpi_cards['active_orders_today'] = {
            'label': _('Активных заказов сегодня'),
            'value': active_orders_today_count,
        }

        # Всего активных подписок
        total_active_subscriptions_count = base_orders_qs.filter(
            order_type=Order.TYPE_SUBSCRIPTION, 
            status=Order.STATUS_CONFIRMED, 
            subscription_end_date__gte=today_date # Подписка еще не закончилась
        ).count()
        kpi_cards['total_active_subscriptions'] = {
            'label': _('Всего активных подписок'),
            'value': total_active_subscriptions_count,
        }
        
        # Дополнительные KPI (если глобальный вид)
        if is_global_view_active:
            if User: 
                kpi_cards['new_users_this_month'] = {
                    'label': _('Новых пользователей (Этот месяц)'),
                    'value': User.objects.filter(date_joined__gte=start_of_this_month).count()
                }
            if University:
                kpi_cards['total_active_universities'] = {
                    'label': _('Активных университетов'),
                    'value': University.objects.filter(is_active=True).count()
                }
            if Facility:
                kpi_cards['total_active_facilities_platform'] = {
                    'label': _('Активных объектов (Платформа)'),
                    'value': Facility.objects.filter(is_active=True, university__is_active=True).count()
                }
            if User:
                kpi_cards['total_active_users_platform'] = {
                    'label': _('Всего активных пользователей'),
                    'value': User.objects.filter(is_active=True).count()
                }
        
        # KPI для конкретного ВУЗа
        if target_university_for_stats:
            if Facility:
                kpi_cards['total_facilities_in_uni'] = {
                    'label': _('Объектов в ВУЗе'),
                    'value': Facility.objects.filter(university=target_university_for_stats, is_active=True).count()
                }
            # Можно добавить другие KPI для ВУЗа
        
        context['stats_kpi_cards'] = kpi_cards
        # --------------------
        
        # Статистика по ВУЗам для таблицы (только для суперюзера и глобального вида)
        if is_global_view_active and University and context.get('universities_list_for_select'):
            universities_stats_list = []
            for uni_in_loop in context['universities_list_for_select']: # Изменил имя переменной цикла
                uni_base_orders_loop = Order.objects.filter(facility__university=uni_in_loop)
                uni_successful_orders_loop = uni_base_orders_loop.filter(status__in=successful_statuses)
                
                uni_revenue_month_loop = uni_successful_orders_loop.filter(created_at__date__gte=start_of_this_month, created_at__date__lte=today_date).aggregate(total=Coalesce(Sum('total_price'), Decimal(0)))['total']
                uni_orders_month_count_loop = uni_successful_orders_loop.filter(created_at__date__gte=start_of_this_month, created_at__date__lte=today_date).count()
                
                uni_total_revenue_ever_loop = uni_successful_orders_loop.aggregate(total=Coalesce(Sum('total_price'), Decimal(0)))['total']
                uni_total_orders_ever_loop = uni_successful_orders_loop.count()
                
                universities_stats_list.append({
                    'id': uni_in_loop.id, 
                    'name': uni_in_loop.name, 
                    'short_name': uni_in_loop.short_name,
                    'month_orders': uni_orders_month_count_loop, 
                    'month_revenue': f"{uni_revenue_month_loop:,}".replace(",", " "),
                    'total_orders_ever': uni_total_orders_ever_loop, 
                    'total_revenue_ever': f"{uni_total_revenue_ever_loop:,}".replace(",", " "),
                    'active_facilities': Facility.objects.filter(university=uni_in_loop, is_active=True).count() if Facility else 0,
                })
            context['universities_stats_table'] = universities_stats_list
            
    return render(request, 'admin/dashboard/statistics_dashboard.html', context)

# ... (остальной код: get_user_dashboard_stats_api, export_orders_to_excel) ...
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_dashboard_stats_api(request):
    user = request.user
    today = timezone.localdate()
    now_time_local = timezone.localtime(timezone.now()).time()

    if not Order: 
        logger.error("get_user_dashboard_stats_api: Order model not available.")
        return Response({"error": "Order model not available for statistics."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    q_user_bookings_active = Q(order_type__in=[Order.TYPE_SLOT_BOOKING, Order.TYPE_ENTRY_FEE]) & \
                             Q(status=Order.STATUS_CONFIRMED) & \
                             (Q(booking_date__gt=today) | (Q(booking_date=today) & ((Q(order_type=Order.TYPE_ENTRY_FEE) & Q(facility__close_time__gt=now_time_local)) | (Q(order_type=Order.TYPE_SLOT_BOOKING) & Q(facility__close_time__gt=now_time_local) )))) # Уточнили для слотов
    q_user_subscriptions_active = Q(order_type=Order.TYPE_SUBSCRIPTION) & \
                                  Q(status=Order.STATUS_CONFIRMED) & \
                                  Q(subscription_end_date__gte=today) & \
                                  Q(subscription_start_date__lte=today)
                                  
    active_orders_count = Order.objects.filter(user=user).filter(q_user_bookings_active | q_user_subscriptions_active).count()
    total_orders_count = Order.objects.filter(user=user, status__in=[Order.STATUS_CONFIRMED, Order.STATUS_COMPLETED]).count()
    active_subscriptions_count = Order.objects.filter(user=user, order_type=Order.TYPE_SUBSCRIPTION, status=Order.STATUS_CONFIRMED, subscription_end_date__gte=today, subscription_start_date__lte=today).count()

    stats_data = {
        'active_orders_count': active_orders_count,
        'total_orders_count': total_orders_count,
        'active_subscriptions_count': active_subscriptions_count,
    }
    return Response(stats_data, status=status.HTTP_200_OK)

def staff_user_check(user):
    return user.is_staff

@user_passes_test(staff_user_check)
def export_orders_to_excel(request):
    user = request.user
    admin_university_obj = get_admin_university_details(user)
    target_university_for_export = None # Изменил имя переменной для ясности
    filename_suffix = "global"

    if user.is_superuser:
        uni_id_str = request.GET.get('university_id')
        if uni_id_str and uni_id_str != "global" and University: # Проверяем, что это не "global"
            try: 
                target_university_for_export = University.objects.get(pk=int(uni_id_str))
                filename_suffix = target_university_for_export.short_name or str(target_university_for_export.id)
            except (University.DoesNotExist, ValueError, TypeError): 
                logger.warning(f"Excel Export: Superuser provided invalid university_id='{uni_id_str}'. Exporting global data.")
                target_university_for_export = None
    elif admin_university_obj and admin_university_obj != -1:
        target_university_for_export = admin_university_obj
        filename_suffix = target_university_for_export.short_name or str(target_university_for_export.id)
    # Если не суперюзер и не админ ВУЗа (admin_university_obj == -1), то экспорт не разрешен
    elif not user.is_superuser and (not admin_university_obj or admin_university_obj == -1):
        logger.warning(f"User {user.email} attempted to export Excel without sufficient permissions.")
        return HttpResponse("Unauthorized", status=403)

    queryset = Order.objects.all()
    if target_university_for_export:
        queryset = queryset.filter(facility__university=target_university_for_export)
    
    orders_to_export = queryset.select_related(
        'user', 'facility', 'facility__university'
    ).order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = gettext("Заказы") 

    columns = [
        gettext("Код Заказа"), gettext("Пользователь (Имя)"), gettext("Пользователь (Email)"), 
        gettext("Объект"), gettext("Университет"), gettext("Тип Заказа"), gettext("Статус"), 
        gettext("Дата/Период"), gettext("Детали (Слоты/Время)"),
        gettext("Сумма (сум)"), gettext("Дата Создания")
    ]
    ws.append(columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell_idx, cell in enumerate(ws[1]): # Используем enumerate для индекса колонки
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Устанавливаем ширину колонки (пример)
        column_letter = chr(65 + cell_idx) # A, B, C...
        if column_letter == 'A': ws.column_dimensions[column_letter].width = 20 # Код заказа
        elif column_letter == 'B': ws.column_dimensions[column_letter].width = 25 # Имя
        elif column_letter == 'C': ws.column_dimensions[column_letter].width = 30 # Email
        elif column_letter == 'D': ws.column_dimensions[column_letter].width = 30 # Объект
        elif column_letter == 'E': ws.column_dimensions[column_letter].width = 20 # ВУЗ
        elif column_letter == 'F': ws.column_dimensions[column_letter].width = 20 # Тип
        elif column_letter == 'G': ws.column_dimensions[column_letter].width = 18 # Статус
        elif column_letter == 'H': ws.column_dimensions[column_letter].width = 20 # Дата/Период
        elif column_letter == 'I': ws.column_dimensions[column_letter].width = 35 # Детали
        elif column_letter == 'J': ws.column_dimensions[column_letter].width = 15 # Сумма
        elif column_letter == 'K': ws.column_dimensions[column_letter].width = 18 # Дата создания


    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for order in orders_to_export:
        user_email = order.user.email if order.user else "-"
        user_name = order.user.get_full_name() if order.user else "-"
        facility_name = order.facility.name if order.facility else "-"
        uni_name = order.facility.university.short_name or order.facility.university.name if order.facility and order.facility.university else "-"
        
        order_dates = "-"
        if order.order_type == Order.TYPE_ENTRY_FEE or order.order_type == Order.TYPE_SLOT_BOOKING:
            order_dates = order.booking_date.strftime('%d.%m.%Y') if order.booking_date else '-'
        elif order.order_type == Order.TYPE_SUBSCRIPTION:
            start = order.subscription_start_date.strftime('%d.%m.%Y') if order.subscription_start_date else '?'
            end = order.subscription_end_date.strftime('%d.%m.%Y') if order.subscription_end_date else '?'
            order_dates = f"{start} - {end}"
        
        order_specifics = "-"
        if order.order_type == Order.TYPE_SLOT_BOOKING:
            order_specifics = order.get_ordered_slots_display() or "-"
        elif order.order_type == Order.TYPE_SUBSCRIPTION:
            days_str_list = order.get_parsed_days_of_week_display() 
            days_str = ", ".join(days_str_list) if days_str_list else '-'
            times_list = order.get_parsed_subscription_times()
            times_str_formatted = "-"
            if times_list:
                duration = order.duration_per_slot_hours or 1
                formatted_times_list = []
                for t_str_item in times_list:
                    try:
                        start_time_obj_item = time.fromisoformat(t_str_item)
                        end_time_obj_item = (datetime.combine(dt_date.min, start_time_obj_item) + timedelta(hours=duration)).time()
                        formatted_times_list.append(f"{start_time_obj_item.strftime('%H:%M')}-{end_time_obj_item.strftime('%H:%M')}")
                    except ValueError: formatted_times_list.append(t_str_item)
                times_str_formatted = ", ".join(formatted_times_list)
            order_specifics = f"{gettext('Дни')}: {days_str}; {gettext('Время')}: {times_str_formatted or '-'}"
        
        created_at_naive = None
        if order.created_at:
            aware_dt = timezone.localtime(order.created_at) # Конвертируем в локальное время
            created_at_naive = aware_dt.replace(tzinfo=None) # Убираем tzinfo для openpyxl

        row_data = [
            order.order_code, user_name, user_email, facility_name, uni_name,
            order.get_order_type_display(), order.get_status_display(),
            order_dates, order_specifics,
            order.total_price, 
            created_at_naive # Используем naive datetime
        ]
        ws.append(row_data)
    
    for row_idx, row_val in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)), 1):
        for cell in row_val:
            cell.border = thin_border
            if row_idx > 1: # Пропускаем заголовки
                if isinstance(cell.value, (int, float, Decimal)):
                     cell.alignment = Alignment(horizontal='right')
                     if columns[cell.column -1] == gettext("Сумма (сум)"): # Проверяем по заголовку
                         cell.number_format = '#,##0" {}"'.format(gettext('сум'))
                elif isinstance(cell.value, datetime): 
                    cell.number_format = 'DD.MM.YYYY HH:MM'
                elif isinstance(cell.value, dt_date): 
                    cell.number_format = 'DD.MM.YYYY'
            elif row_idx == 1: # Заголовки
                 cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename_date_part = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    final_filename = f"unisport_orders_{filename_suffix}_{filename_date_part}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{final_filename}"'
    
    try: wb.save(response)
    except Exception as e:
        logger.error(f"Error saving workbook to response: {e}", exc_info=True)
        return HttpResponse("Error generating Excel file.", status=500)
    return response
# --- END OF FULL FILE ---